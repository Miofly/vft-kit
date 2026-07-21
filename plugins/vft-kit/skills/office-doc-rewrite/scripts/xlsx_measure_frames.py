# -*- coding: utf-8 -*-
"""
量出 xlsx 里每张图片 twoCellAnchor 显示框的真实像素尺寸与长宽比。
新图要 padding 到这个比例才不会被 Excel 拉伸变形。
用法: $OFFICE_PY xlsx_measure_frames.py <file.xlsx>
输出每张 image 的 框宽x框高px 和 比例,以及它挂在哪个 sheet。
"""
import sys, re, zipfile
import openpyxl
from openpyxl.utils import get_column_letter

def colwidth_px(ws, col_idx):
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    w = dim.width if (dim and dim.width) else (ws.sheet_format.defaultColWidth or 8.43)
    return w * 7 + 5

def rowht_px(ws, row_idx):
    dim = ws.row_dimensions.get(row_idx)
    h = dim.height if (dim and dim.height) else (ws.sheet_format.defaultRowHeight or 15)
    return h * 96 / 72

def main():
    xlsx = sys.argv[1]
    z = zipfile.ZipFile(xlsx)
    wb = openpyxl.load_workbook(xlsx)

    # sheet.xml → drawing → image 映射
    # 找每个 sheet 的 rels 指向的 drawing
    names = z.namelist()
    # sheetN.xml → 工作表名
    wbxml = z.read('xl/workbook.xml').decode('utf-8')
    rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    rid2file = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/[^"]+)"', rels))
    name2file = {}
    for nm, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml):
        name2file[nm] = rid2file.get(rid)

    for sheet_name in wb.sheetnames:
        sfile = name2file.get(sheet_name)
        if not sfile:
            continue
        rels_path = f"xl/worksheets/_rels/{sfile.split('/')[-1]}.rels"
        if rels_path not in names:
            continue
        srels = z.read(rels_path).decode('utf-8')
        dm = re.search(r'Target="\.\./(drawings/[^"]+)"', srels)
        if not dm:
            continue
        dpath = 'xl/' + dm.group(1)
        drawing = z.read(dpath).decode('utf-8')
        drels = z.read(f"xl/drawings/_rels/{dpath.split('/')[-1]}.rels").decode('utf-8')
        rid2img = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"', drels))
        ws = wb[sheet_name]
        for a in re.findall(r'<xdr:twoCellAnchor.*?</xdr:twoCellAnchor>', drawing, re.S):
            rid = re.search(r'r:embed="(rId\d+)"', a)
            fr = re.search(r'<xdr:from><xdr:col>(\d+)</xdr:col><xdr:colOff>(\d+)</xdr:colOff><xdr:row>(\d+)</xdr:row><xdr:rowOff>(\d+)</xdr:rowOff>', a)
            to = re.search(r'<xdr:to><xdr:col>(\d+)</xdr:col><xdr:colOff>(\d+)</xdr:colOff><xdr:row>(\d+)</xdr:row><xdr:rowOff>(\d+)</xdr:rowOff>', a)
            if not (rid and fr and to):
                continue
            img = rid2img.get(rid.group(1), '?')
            fc, fco, frow, fro = map(int, fr.groups())
            tc, tco, trow, tro = map(int, to.groups())
            wpx = sum(colwidth_px(ws, c + 1) for c in range(fc, tc)) + (tco - fco) / 9525
            hpx = sum(rowht_px(ws, r + 1) for r in range(frow, trow)) + (tro - fro) / 9525
            ratio = wpx / hpx if hpx else 0
            print(f"{img} [{sheet_name}]: 框≈{int(wpx)}x{int(hpx)}px  比{ratio:.2f}")

if __name__ == '__main__':
    main()
